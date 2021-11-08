## Descarga de Reporte Bretti 
## By Equipo BI DoBetter
## Funcion: Descarga el reporte de GPS para el Estado de GPS DIARIO DE UNILEVER
## Fecha 29-10-2021 , Santiago de Chile.

###########################
#  Bloque de importacion  #
###########################

from openpyxl.worksheet.table import Table
from typing import Container
import shutil
from openpyxl.worksheet.table import TableStyleInfo
from pandas.core.frame import DataFrame
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.common.action_chains import ActionChains
import os
from datetime import date
from datetime import datetime
from datetime import timedelta
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook
import warnings

###########################
#  Bloque de definiciones #
###########################

## Fecha debe estar en formato "dd-mm-yyyy" 
Hoy= date.today()
## Si estoy en el lunes y quiero el del viernes agrego dias al delta
Ndias = 1
fechaInicial = date.today()- timedelta( days= Ndias)
fechaFinal= fechaInicial

fechainicial_formato = fechaInicial.strftime("%d-%m-%Y")
fechafinal_formato= fechaFinal.strftime("%d-%m-%Y")
hoy_formato= Hoy.strftime("%d-%m-%Y")

source = r"C:\Users\bsaaa\Dropbox (Do Better)\PC\Downloads"
usuario_ruta = "gusta"
destiny= r"C:\Users\bsaaa\Dropbox (Do Better)\PC\Desktop\Reportes\Unilever\Estado de GPS"
destiny_reumido = r"C:\Users\bsaaa\Dropbox (Do Better)\PC\Desktop\Reportes\Unilever\Estado de GPS\Noviembre\resumido"

File = "Listado de viajes.xlsx"
File_resumido = "Listado de viajes_resumido.xlsx"

renombre= "Estado GPS "+fechainicial_formato+", Bajado el "+hoy_formato+".xlsx"
renombre_resumido= "Estado GPS RESUMIDO "+fechainicial_formato+", Bajado el "+hoy_formato+".xlsx"


ruta = "C:\driver chrome\chromedriver"

UserNameWing= "Andrea@dobetter.cl"
PassWing = "Andrea2020"

## Columnas que usa el archivo resumido
col1 = "Código Viaje"
col2 = "Nombre Transportista"
col3= "Patente Móvil"
col4 = "Estado del GPS"

rutaUnilever= r"C:\Users\bsaaa\Dropbox (Do Better)\Do Better's shared workspace\Asesorías\GO\Gestion Administrativo GO\Administrativo\Reportes\Unilever\Estado GPS Transportes UL"
filenameUnilever= "Listado de viajes Unilever_automatico.xlsx"


################################################0
################################################

##Bloque de procesos
    
def GPS_Diario():
    ## Funcion que Ira a la pagina y obtendra los datos en PORTAPAPELES, ya que al descargar la info 
    ## no viene en un formato real de xlsx o csv

    ## Flujo de trabajo: 1- Encontrar el elemento. 2- Interactuar con el elemento
    ## Se agregan try y except para controlar y monitorear errores de ejecucion

    # ingreso pagina
    try:
        driver= webdriver.Chrome(executable_path=ruta)
        driver.get("https://suite.wing.cl/web/core/inicio_sesion.php")
        driver.maximize_window()
    except:
        print("Error tipo 0")

    # Ingresar usuario
    try:
        input_user = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="page-container"]/div/div[2]/div[2]/form/div[1]/input[1]')))
        
        # Forma antigua, menos optima
        #inputUser = driver.find_element_by_xpath('//*[@id="page-container"]/div/div[2]/div[2]/form/div[1]/input[1]')
        
        input_user.send_keys(UserNameWing)
    except:
        print("Erro tipo 1")
       ## LLAMAR A LA FUNCION 

    # Contrasena
    try:
        contra = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'password')))        
        #contra = driver.find_element_by_name('password')
        contra.send_keys(PassWing)
        contra.send_keys(Keys.ENTER)
    except:
        print("Error tipo 2")

    ## Cambio de pestanha , al tema de logistics
    try:
        ## Entro 
        sleep(2)
        driver.get("https://suite.wing.cl/web/core/index.php?id_app=5&id_f=18")
        sleep(3)
        boton1= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="side-menu"]/li[7]/a')))
        #boton1= driver.find_element_by_xpath('//*[@id="side-menu"]/li[7]/a')
        boton1.click()
    except:
        print("Error tipo 3")

    ## Hasta aca son iguales

    ## Elegir reporte 
    try:
        eleccion_reporte= WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.LINK_TEXT,'Informe Base Viajes Unilever')))
        #boton2= driver.find_element_by_xpath('//*[@id="side-menu"]/li[7]/ul[3]/li/a')
        eleccion_reporte.click()
        sleep(3)
    except:
        print("Error tipo 4")

    try:
        filtroFechaInicial= WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="fecha_inicio"]' )))
        #filtroFechaInicial=driver.find_element_by_xpath('//*[@id="fecha_inicio"]')

        ## Primero se borra lo que esta escrito
        ## Iterador
        i = 0
        while i < 50:
            filtroFechaInicial.send_keys(Keys.BACK_SPACE)
            i= i +1 

        sleep(2)

        #filtroFechaInicial= driver.find_element_by_xpath('//*[@id="fecha_inicio"]')
        filtroFechaInicial.send_keys(fechainicial_formato)

        fechaFinal= WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH , '//*[@id="fecha_fin"]')))
        #fechaFinal= driver.find_element_by_xpath('//*[@id="fecha_fin"]')

        j = 0

        while j < 50:
            fechaFinal.send_keys(Keys.BACK_SPACE)
            j= j +1 

        fechaFinal.send_keys(fechainicial_formato)
        fechaFinal.send_keys(Keys.ENTER)
    
    except:
        print("Error tipo 5")
    sleep(3)

    try:
        botonVerDatos=WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="filtros_movil"]/div/div[8]/button' )))
        #botonVerDatos = driver.find_element_by_xpath('//*[@id="filtros_movil"]/div/div[8]/button')
        botonVerDatos.click()
    except:
        print("Error tipo 6")

    try:
        sleep(5)
        botoncopiar= WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,'//*[@id="DataTables_Table_0_wrapper"]/div[1]/div/a[1]' )))
        #botoncopiar= driver.find_element_by_xpath('//*[@id="DataTables_Table_0_wrapper"]/div[1]/div/a[1]')
        botoncopiar.click()
    
    except:
        print("Error tipo 7")


    sleep(3)
    #driver.quit()
    print("Reporte guardado en el PortaPapeles exitosamente")


def Exportacion_GPS():
    ## Expotacion a Pandas desde el PortaPapeles (CLIPBOARD)
    sleep(3)
    df2 = pd.read_clipboard(error_bad_lines=False)
    print(df2.head())

    ## Cambio la ruta
    os.chdir(source)

    ## Renombrar archivo ya sea CSV o Excel
    #df2.to_csv("FINALMENTELISTO.csv")
    df2.to_excel(File)
    ## Crear archivo filtrado 

    #df_No_Asignadas = df2[df2["Estado Viaje"]=="No asignado"]
    #print(df_No_Asignadas.head())
    ## Comienza a trabajar con OS 

    df_resumido = df2[[col1,col2,col3,col4]]
    #print(df_resumido.head())
    #print(df_resumido.info())

    df_resumido.to_excel(File_resumido)

    sleep(5)

    ## Mover archivo 
    shutil.move(File , destiny)
    shutil.move(File_resumido, destiny_reumido)

    ## Renombrar

    os.chdir(destiny)
    os.rename(File , renombre)


    os.chdir(destiny_reumido)
    os.rename(File_resumido , renombre_resumido)

    ## Ahora voy a tomar los xlsx y mover la info a los definitivos 
    warnings.simplefilter("ignore")

    ## Ir a la ruta del archivo filtrado
    os.chdir(destiny_reumido)
    wb1 = xl.load_workbook(renombre_resumido)
    ws1 = wb1.worksheets[0]

    ## ir a la ruta de destino 
    os.chdir(rutaUnilever)

    ## Abrir arhcivo
    wb2= xl.load_workbook(filenameUnilever)
    ws2= wb2.active


    # calculate total number of rows and 
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column
    max_row_2 = ws2.max_row

    ## Para una columna en especifico
    for i in range(2, mr +1):
        c=ws1.cell(row=i,column=2)
        ws2.cell(row=i+max_row_2-1, column=3).value=c.value
    ############################################################
    for i in range(2, mr +1):
        c=ws1.cell(row=i,column=3)
        ws2.cell(row=i+max_row_2-1, column=4).value=c.value  

    for i in range(2, mr +1):
        c=ws1.cell(row=i,column=4)
        ws2.cell(row=i+max_row_2-1, column=5).value=c.value

    for i in range(2, mr +1):
        c=ws1.cell(row=i,column=5)
        ws2.cell(row=i+max_row_2-1, column=6).value=c.value
    
    for i in range(2, mr +1):
        c=ws1.cell(row=i,column=5)
        ws2.cell(row=i+max_row_2-1, column=7).value=c.value

    ## Ahora id y fecha 

    ##Fecha
    for i in range(2, mr +1):
        ws2.cell(row=i+max_row_2-1, column=1).value= fechainicial_formato

    ##ID 
    ##Volver dinamico al ID actual
    id_Actual = max_row_2 + 30104  + 7065

    #for i in range(2, mr +1):
    #    iterador = 1
    #    ws2.cell(row=i+max_row_2-1, column=2).value= id_Actual + iterador
    #    interador = iterador + 1
    #    print(iterador)


    # ID con while 

    iterador= 1
    j = 2
    while j < mr+1:
        ws2.cell(row=j+max_row_2-1, column=2).value= id_Actual + iterador
        iterador = iterador+1
        j = j +1

    ## Columna H
    for i in range(2, mr +1):
        ws2.cell(row=i+max_row_2-1, column=8).value= "VERDADERO"

    ## Dar Formato tabla
    ## Primero eliminar tabla existente

    del ws2.tables["Viajes_UL"]

    # define a table style
    referencia = "A1:H"+str(max_row_2 + mr - 1)

    mediumStyle = xl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                        showRowStripes=True)
    # create a table
    table = xl.worksheet.table.Table(ref= referencia,
                                        displayName='Viajes_UL',
                                        tableStyleInfo=mediumStyle)
    # add the table to the worksheet
    ws2.add_table(table)

    # save the workbook file
    #wb2.save('fruit.xlsx')

    print(ws2.tables.items())

    wb2.save(filenameUnilever)
    ## FIN 
    print("fin")


GPS_Diario()
#Exportacion_GPS()

'''

# copying the cell values from source 
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
  
        # writing the read value to destination excel file
        ws2.cell(row = i, column = j).value = c.value
  
# saving the destination excel file
wb2.save(str(filename1))

'''

